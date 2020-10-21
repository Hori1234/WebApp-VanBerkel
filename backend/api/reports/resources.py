from flask import safe_join, send_file, send_from_directory
from flask_smorest import abort
from flask.views import MethodView
from sqlalchemy import func, and_
from . import bp
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
import time
from backend.app import db
from backend.models.orders import Order, OrderSheet
from backend.extensions import roles_required
from backend.api.reports.schemas import ReportSchema
import io


@bp.route('/firstrides/<sheet_id_or_latest>')
class FirstRides(MethodView):

    @bp.response(ReportSchema, code=200)
    @bp.alt_response("BAD_REQUEST", code=400)
    @roles_required('planner', 'administrator')
    def get(self, sheet_id_or_latest):
        """
        Get a workbook containing a first rides report from an order sheet.

        In case 'sheet_id_or_latest is 'latest', the most recently uploaded
        order sheet will be used.
        """
        try:
            # Try to parse the sheet_id to an int and get the order sheet
            order_sheet = OrderSheet.query \
                .get_or_404(int(sheet_id_or_latest),
                            description="Order sheet not found")
        except ValueError:
            # The sheet_id is not an integer, so check if it is `latest`
            if sheet_id_or_latest == 'latest':
                # Get the most recently uploaded sheet
                order_sheet = OrderSheet.query\
                    .order_by(OrderSheet.upload_date.desc())\
                    .first_or_404()
            else:
                # Can't understand which sheet is requested
                return abort(404, message='Order sheet not found')

        sheet_id = order_sheet.id

        subq = db.session.query(
            Order.truck_id,
            func.min(Order.departure_time).label('mintime')
        ).group_by(Order.truck_id).filter(Order.sheet_id == sheet_id) \
            .subquery()

        first_orders = db.session.query(Order).join(
            subq,
            and_(
                Order.truck_id == subq.c.truck_id,
                Order.departure_time == subq.c.mintime
            )
        ).all()

        book = Workbook()
        sheet = book.active
        now = time.strftime("%d %b %Y %X")
        now_save = time.strftime("%Y-%m-%d-%H-%M-%S")

        sheet['A1'] = now
        sheet['C4'] = 'Sno'
        sheet['D4'] = 'Driver Name'
        sheet['E4'] = 'Truck ID'
        sheet['F4'] = 'Terminal'
        sheet['G4'] = 'chassis'
        sheet['H4'] = 'Starting Time'
        sheet['I4'] = 'Delivery Deadline'
        sheet['J4'] = 'Customer'
        sheet['K4'] = 'Container No.'
        sheet['L4'] = 'City'
        sheet['M4'] = 'Container Type'
        sheet['N4'] = 'Shipping company'
        sheet['O4'] = 'Remarks'

        yellowFill = PatternFill(start_color='FCE205',
                                 end_color='FFD300',
                                 fill_type='solid')
        sheet['C4'].fill = yellowFill
        sheet['D4'].fill = yellowFill
        sheet['E4'].fill = yellowFill
        sheet['F4'].fill = yellowFill
        sheet['G4'].fill = yellowFill
        sheet['H4'].fill = yellowFill
        sheet['I4'].fill = yellowFill
        sheet['J4'].fill = yellowFill
        sheet['K4'].fill = yellowFill
        sheet['L4'].fill = yellowFill
        sheet['M4'].fill = yellowFill
        sheet['N4'].fill = yellowFill
        sheet['O4'].fill = yellowFill

        sheet['C4'].font = Font(bold=True)
        sheet['D4'].font = Font(bold=True)
        sheet['E4'].font = Font(bold=True)
        sheet['F4'].font = Font(bold=True)
        sheet['G4'].font = Font(bold=True)
        sheet['H4'].font = Font(bold=True)
        sheet['I4'].font = Font(bold=True)
        sheet['J4'].font = Font(bold=True)
        sheet['K4'].font = Font(bold=True)
        sheet['L4'].font = Font(bold=True)
        sheet['M4'].font = Font(bold=True)
        sheet['N4'].font = Font(bold=True)
        sheet['O4'].font = Font(bold=True)

        sheet.column_dimensions['D'].width = 20
        sheet.column_dimensions['H'].width = 15
        sheet.column_dimensions['I'].width = 15
        sheet.column_dimensions['N'].width = 20
        sheet.column_dimensions['O'].width = 40

        count = 4
        for order in first_orders:
            count = count + 1

            sheet.cell(row=count, column=3).value = \
                order.truck.s_number  # s number
            sheet.cell(row=count, column=4).value = \
                order.truck.others.get('Driver', '')   # driver
            sheet.cell(row=count, column=5).value = \
                order.truck.truck_id  # truck id
            sheet.cell(row=count, column=6).value = \
                order.inl_terminal  # terminal
            sheet.cell(row=count, column=7).value = \
                ''  # chassis?
            sheet.cell(row=count, column=8).value = \
                order.departure_time  # dep time
            sheet.cell(row=count, column=9).value = \
                order.delivery_deadline  # deadline
            sheet.cell(row=count, column=10).value = \
                order.others.get('Client', '')  # client
            sheet.cell(row=count, column=11).value = \
                order.others.get('Container', '')  # container number
            sheet.cell(row=count, column=12).value = \
                order.others.get('City', '')  # city
            sheet.cell(row=count, column=13).value = \
                order.others.get('Unit type', '')  # container type
            sheet.cell(row=count, column=14).value = \
                order.others.get('Ship. comp.', '')  # shipping company
            sheet.cell(row=count, column=15).value = \
                order.truck.others.get('Remarks', '')  # remarks

        filename = 'first-rides-' + now_save + '.xlsx'
        file = io.BytesIO()
        book.save(file)
        file.seek(0)

        try:
            return send_file(
                file,
                attachment_filename=filename,
                as_attachment=True
                ), 200
        except FileNotFoundError:
            abort(404)

        return '', 200
